import requests
import time
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os

class AccelerationDataCollector:
    """가속도 데이터 수집 및 엑셀 저장"""
    
    def __init__(self,
                 top_n=20,
                 check_interval=10,
                 duration_minutes=None):
        
        self.top_n = top_n
        self.check_interval = check_interval
        self.duration_minutes = duration_minutes
        self.base_url = "https://api.upbit.com/v1"
        
        # 데이터 저장
        self.data_history = []  # 모든 기록
        self.prices_prev = {}  # 이전 시점 가격
        self.velocities_prev = {}  # 이전 시점 속도

        # 최대/최소 가속도 추적
        self.max_accel_record = None  # 최대 가속도 기록
        self.min_accel_record = None  # 최소 가속도 기록
        self.max_accel_tracking = []  # 최대 가속도 이후 가격 변화 추적

        # 모니터링 대상 코인
        self.target_markets = []
    
    def get_top_gainers(self):
        """전일대비 상승률 상위 20개 조회"""
        url = f"{self.base_url}/ticker"
        
        try:
            # 모든 원화 마켓 조회
            markets_url = f"{self.base_url}/market/all"
            markets_response = requests.get(markets_url, params={"isDetails": "false"})
            markets = markets_response.json()
            krw_markets = [m['market'] for m in markets if m['market'].startswith('KRW-')]
            
            # 전체 티커 조회
            params = {"markets": ",".join(krw_markets)}
            response = requests.get(url, params=params)
            tickers = response.json()
            
            # 상승률 순으로 정렬
            sorted_tickers = sorted(tickers, 
                                   key=lambda x: x['signed_change_rate'], 
                                   reverse=True)
            
            # 상위 N개
            top_gainers = sorted_tickers[:self.top_n]
            
            print(f"\n{'='*80}")
            print(f"📊 전일대비 상승률 상위 {self.top_n}개 종목")
            print(f"{'='*80}")
            print(f"{'순위':<4} {'종목':<12} {'현재가':<15} {'전일대비':<12}")
            print("-" * 80)
            
            for i, ticker in enumerate(top_gainers, 1):
                change_pct = ticker['signed_change_rate'] * 100
                print(f"{i:<4} {ticker['market']:<12} "
                      f"{ticker['trade_price']:>12,.2f}원 "
                      f"{change_pct:>+9.2f}%")
            
            print(f"{'='*80}\n")
            
            return [t['market'] for t in top_gainers]
            
        except Exception as e:
            print(f"❌ 조회 실패: {e}")
            return []
    
    def get_target_tickers(self):
        """모니터링 대상 코인들의 현재가 조회"""
        url = f"{self.base_url}/ticker"
        params = {"markets": ",".join(self.target_markets)}
        
        try:
            response = requests.get(url, params=params)
            return response.json()
        except Exception as e:
            print(f"❌ 티커 조회 실패: {e}")
            return None
    
    def calculate_metrics(self, tickers, timestamp):
        """속도와 가속도 계산"""
        records = []

        for ticker in tickers:
            market = ticker['market']
            current_price = ticker['trade_price']
            change_rate_24h = ticker['signed_change_rate'] * 100

            # 속도 계산 (이전 시점 가격 필요)
            velocity = None
            acceleration = None

            if market in self.prices_prev:
                price_prev = self.prices_prev[market]

                # 속도 = (현재가 - 이전가) / 이전가 × 100
                velocity = ((current_price - price_prev) / price_prev) * 100

                # 가속도 계산 (이전 시점 속도 필요)
                if market in self.velocities_prev:
                    velocity_prev = self.velocities_prev[market]

                    # 가속도 = 현재속도 - 이전속도
                    acceleration = velocity - velocity_prev

            # 기록 저장
            record = {
                '시간': timestamp,
                '종목': market,
                '현재가': current_price,
                '전일대비(%)': change_rate_24h,
                f'속도_{self.check_interval}초(%)': velocity if velocity is not None else 0,
                '가속도(%p)': acceleration if acceleration is not None else 0,
                '이전가격': self.prices_prev.get(market, current_price)
            }

            records.append(record)

            # 최대/최소 가속도 업데이트
            if acceleration is not None:
                if self.max_accel_record is None or acceleration > self.max_accel_record['가속도(%p)']:
                    self.max_accel_record = record.copy()
                    print(f"\n🔥 새로운 최대 가속도 발견!")
                    print(f"   종목: {market} | 가속도: {acceleration:+.4f}%p | 가격: {current_price:,.0f}원")

                if self.min_accel_record is None or acceleration < self.min_accel_record['가속도(%p)']:
                    self.min_accel_record = record.copy()
                    print(f"\n❄️  새로운 최소 가속도 발견!")
                    print(f"   종목: {market} | 가속도: {acceleration:+.4f}%p | 가격: {current_price:,.0f}원")

            # 다음 계산을 위해 저장
            self.prices_prev[market] = current_price
            if velocity is not None:
                self.velocities_prev[market] = velocity

        return records
    
    def print_current_status(self, records):
        """현재 상태 출력"""
        print(f"\n{'='*100}")
        print(f"[{records[0]['시간']}] 📊 실시간 모니터링")
        print(f"{'='*100}")

        # 동적으로 속도 컬럼명 가져오기
        velocity_col = [k for k in records[0].keys() if k.startswith('속도_')][0]

        print(f"{'종목':<12} {'현재가':<15} {'전일대비':<12} {'속도':<15} {'가속도':<15}")
        print("-" * 100)

        # 가속도 높은 순으로 정렬
        sorted_records = sorted(records, key=lambda x: x['가속도(%p)'], reverse=True)

        for record in sorted_records[:10]:  # 상위 10개만 출력
            velocity_str = f"{record[velocity_col]:+.4f}%" if record[velocity_col] != 0 else "계산중"
            accel_str = f"{record['가속도(%p)']:+.4f}%p" if record['가속도(%p)'] != 0 else "계산중"

            print(f"{record['종목']:<12} "
                  f"{record['현재가']:>12,.2f}원 "
                  f"{record['전일대비(%)']:>+9.2f}% "
                  f"{velocity_str:>13} "
                  f"{accel_str:>13}")

        print(f"{'='*100}\n")
    
    def save_to_excel(self):
        """엑셀 파일로 저장"""
        if not self.data_history:
            print("❌ 저장할 데이터가 없습니다.")
            return
        
        # DataFrame 생성
        df = pd.DataFrame(self.data_history)
        
        # 파일명 (타임스탬프 포함)
        filename = f"acceleration_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 엑셀 작성
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 전체 데이터 시트
            df.to_excel(writer, sheet_name='전체데이터', index=False)
            
            # 종목별 시트 (각 종목마다 별도 시트)
            for market in self.target_markets:
                market_df = df[df['종목'] == market].copy()
                if not market_df.empty:
                    sheet_name = market.replace('KRW-', '')[:31]  # 엑셀 시트명 길이 제한
                    market_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 요약 통계 시트
            summary_data = []
            # 동적으로 속도 컬럼명 찾기
            velocity_cols = [col for col in df.columns if col.startswith('속도_')]
            velocity_col = velocity_cols[0] if velocity_cols else None

            for market in self.target_markets:
                market_df = df[df['종목'] == market]

                if len(market_df) > 0 and velocity_col:
                    summary = {
                        '종목': market,
                        '평균속도': market_df[velocity_col].mean(),
                        '최대속도': market_df[velocity_col].max(),
                        '최소속도': market_df[velocity_col].min(),
                        '평균가속도': market_df['가속도(%p)'].mean(),
                        '최대가속도': market_df['가속도(%p)'].max(),
                        '최소가속도': market_df['가속도(%p)'].min(),
                        '최종가격': market_df.iloc[-1]['현재가'],
                        '시작가격': market_df.iloc[0]['현재가'],
                        '총변화율(%)': ((market_df.iloc[-1]['현재가'] - market_df.iloc[0]['현재가'])
                                     / market_df.iloc[0]['현재가'] * 100)
                    }
                    summary_data.append(summary)
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='요약통계', index=False)

            # 최대/최소 가속도 분석 시트
            if self.max_accel_record and self.min_accel_record:
                accel_analysis = []

                # 최대 가속도 정보
                max_info = {
                    '구분': '최대 가속도',
                    '종목': self.max_accel_record['종목'],
                    '발생시간': self.max_accel_record['시간'],
                    '가속도(%p)': self.max_accel_record['가속도(%p)'],
                    '발생시가격': self.max_accel_record['현재가'],
                    '전일대비(%)': self.max_accel_record['전일대비(%)']
                }
                accel_analysis.append(max_info)

                # 최소 가속도 정보
                min_info = {
                    '구분': '최소 가속도',
                    '종목': self.min_accel_record['종목'],
                    '발생시간': self.min_accel_record['시간'],
                    '가속도(%p)': self.min_accel_record['가속도(%p)'],
                    '발생시가격': self.min_accel_record['현재가'],
                    '전일대비(%)': self.min_accel_record['전일대비(%)']
                }
                accel_analysis.append(min_info)

                accel_df = pd.DataFrame(accel_analysis)
                accel_df.to_excel(writer, sheet_name='최대최소가속도', index=False)

            # 최대 가속도 종목 추적 시트
            if self.max_accel_tracking:
                tracking_df = pd.DataFrame(self.max_accel_tracking)
                tracking_df.to_excel(writer, sheet_name='최대가속도종목추적', index=False)

        # 엑셀 스타일 적용
        self.apply_excel_formatting(filename)
        
        print(f"\n{'='*80}")
        print(f"✅ 데이터 저장 완료!")
        print(f"파일명: {filename}")
        print(f"총 데이터: {len(self.data_history)}개 기록")

        # 최대 가속도 종목의 최종 결과 출력
        if self.max_accel_tracking:
            last_tracking = self.max_accel_tracking[-1]
            print(f"\n📊 최대 가속도 종목 분석:")
            print(f"   종목: {last_tracking['종목']}")
            print(f"   최대가속도: {last_tracking['최대가속도']:+.4f}%p")
            print(f"   발생시간: {last_tracking['최대가속도발생시간']}")
            print(f"   발생시가격: {last_tracking['발생시가격']:,.0f}원")
            print(f"   최종가격: {last_tracking['현재가격']:,.0f}원")
            print(f"   가격변화: {last_tracking['가격변화(%)']:+.2f}%")
            print(f"   경과시간: {last_tracking['경과시간(분)']:.1f}분")

        print(f"{'='*80}\n")
    
    def apply_excel_formatting(self, filename):
        """엑셀 서식 적용"""
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        wb = load_workbook(filename)
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # 모든 시트에 적용
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 헤더 서식
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 가속도 양수/음수 색상 (전체데이터 시트)
            if sheet_name == '전체데이터':
                accel_col = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == '가속도(%p)':
                        accel_col = idx
                        break
                
                if accel_col:
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=accel_col)
                        if isinstance(cell.value, (int, float)):
                            if cell.value > 0:
                                cell.fill = green_fill
                            elif cell.value < 0:
                                cell.fill = red_fill
        
        wb.save(filename)
    
    def run(self):
        """메인 실행"""
        print(f"\n{'='*80}")
        print(f"📊 가속도 데이터 수집 시작")
        print(f"{'='*80}")
        print(f"수집 간격: {self.check_interval}초")
        if self.duration_minutes:
            print(f"수집 시간: {self.duration_minutes}분")
        else:
            print(f"수집 시간: 무제한 (Ctrl+C로 종료)")
        print(f"대상 종목: 상위 {self.top_n}개")
        print(f"{'='*80}")
        
        # 상위 종목 선정
        self.target_markets = self.get_top_gainers()
        
        if not self.target_markets:
            print("❌ 대상 종목을 찾을 수 없습니다.")
            return
        
        print(f"✅ 모니터링 시작... (Ctrl+C로 중지)")
        if self.duration_minutes:
            end_time = datetime.fromtimestamp(datetime.now().timestamp() + self.duration_minutes * 60)
            print(f"⏰ 예상 종료 시간: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        
        start_time = time.time()
        iteration = 0
        
        try:
            while True:
                iteration += 1
                elapsed_minutes = (time.time() - start_time) / 60

                # 시간 종료 체크 (duration_minutes가 설정된 경우만)
                if self.duration_minutes and elapsed_minutes >= self.duration_minutes:
                    print(f"\n⏰ 설정 시간({self.duration_minutes}분) 도달. 종료합니다.")
                    break

                # 데이터 수집
                tickers = self.get_target_tickers()
                
                if tickers:
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    records = self.calculate_metrics(tickers, timestamp)

                    # 전체 히스토리에 추가
                    self.data_history.extend(records)

                    # 최대 가속도 종목의 가격 변화 추적
                    if self.max_accel_record:
                        max_accel_market = self.max_accel_record['종목']
                        max_accel_price = self.max_accel_record['현재가']
                        max_accel_time = self.max_accel_record['시간']

                        # 현재 가격 찾기
                        current_record = next((r for r in records if r['종목'] == max_accel_market), None)
                        if current_record:
                            current_price = current_record['현재가']
                            price_change_pct = ((current_price - max_accel_price) / max_accel_price) * 100

                            tracking_record = {
                                '측정시간': timestamp,
                                '최대가속도발생시간': max_accel_time,
                                '종목': max_accel_market,
                                '최대가속도': self.max_accel_record['가속도(%p)'],
                                '발생시가격': max_accel_price,
                                '현재가격': current_price,
                                '가격변화(%)': price_change_pct,
                                '경과시간(분)': (datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S') -
                                               datetime.strptime(max_accel_time, '%Y-%m-%d %H:%M:%S')).total_seconds() / 60
                            }
                            self.max_accel_tracking.append(tracking_record)

                    # 상태 출력
                    self.print_current_status(records)

                    if self.duration_minutes:
                        remaining = self.duration_minutes - elapsed_minutes
                        print(f"⏱️  진행: {iteration}회차 | 경과: {elapsed_minutes:.1f}분 | "
                              f"남은시간: {remaining:.1f}분 | 총 데이터: {len(self.data_history)}개")
                    else:
                        print(f"⏱️  진행: {iteration}회차 | 경과: {elapsed_minutes:.1f}분 | "
                              f"총 데이터: {len(self.data_history)}개")
                
                time.sleep(self.check_interval)
                
        except KeyboardInterrupt:
            print("\n\n🛑 사용자가 중지했습니다.")
        
        # 엑셀 저장
        print("\n💾 데이터를 엑셀로 저장 중...")
        self.save_to_excel()
        
        # 최종 통계
        if self.data_history:
            df = pd.DataFrame(self.data_history)
            
            print(f"\n{'='*80}")
            print(f"📈 수집 완료 통계")
            print(f"{'='*80}")
            print(f"총 수집 시간: {(time.time() - start_time) / 60:.1f}분")
            print(f"총 수집 횟수: {iteration}회")
            print(f"총 데이터 수: {len(self.data_history)}개")
            print(f"모니터링 종목: {len(self.target_markets)}개")
            
            # 가속도 상위/하위
            valid_data = df[df['가속도(%p)'] != 0]
            if not valid_data.empty:
                print(f"\n🔥 최대 가속도:")
                max_accel = valid_data.loc[valid_data['가속도(%p)'].idxmax()]
                print(f"   {max_accel['종목']} | {max_accel['시간']} | {max_accel['가속도(%p)']:+.4f}%p")
                
                print(f"\n❄️  최소 가속도:")
                min_accel = valid_data.loc[valid_data['가속도(%p)'].idxmin()]
                print(f"   {min_accel['종목']} | {min_accel['시간']} | {min_accel['가속도(%p)']:+.4f}%p")
            
            print(f"{'='*80}\n")


# 실행
if __name__ == "__main__":
    collector = AccelerationDataCollector(
        top_n=20,              # 상위 20개 종목
        check_interval=30,     # 30초마다 수집
        duration_minutes=360   # 6시간 (360분) 동안 수집
    )
    
    collector.run()